"""openpyxl-based extraction: cell/formula walk, merges, tables, protection.

openpyxl cannot evaluate formulas — we never ask it to. `data_only=False`
(the default) guarantees we get formula text as authored, not a cached
result. Where openpyxl's object model is known to drop information (x14
data-validation extensions, some print/protection edge cases), the raw XML
reader (xml_reader.py) is used instead or in addition — see scan.py for how
the two are merged per sheet.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

FORMULA_CATEGORY_RULES = [
    ("LOOKUP", re.compile(r"\b(VLOOKUP|HLOOKUP|INDEX|MATCH|XLOOKUP|LOOKUP)\s*\(", re.I)),
    ("CONDITIONAL", re.compile(r"\b(IF|IFS|IFERROR|IFNA|SUMIF|SUMIFS|COUNTIF|COUNTIFS|AVERAGEIF)\s*\(", re.I)),
    ("AGGREGATE", re.compile(r"\b(SUM|SUBTOTAL|AVERAGE|MAX|MIN|COUNT|PRODUCT)\s*\(", re.I)),
    ("TEXT", re.compile(r"\b(CONCATENATE|TEXT|LEFT|RIGHT|MID|TRIM|UPPER|LOWER|SUBSTITUTE|&)", re.I)),
    ("DATE", re.compile(r"\b(TODAY|NOW|DATE|DATEDIF|YEAR|MONTH|DAY|EDATE)\s*\(", re.I)),
    ("ROUNDING", re.compile(r"\b(ROUND|ROUNDUP|ROUNDDOWN|CEILING|FLOOR|TRUNC|INT)\s*\(", re.I)),
    ("VOLATILE", re.compile(r"\b(NOW|TODAY|RAND|RANDBETWEEN|OFFSET|INDIRECT|CELL|INFO)\s*\(", re.I)),
]

CELL_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*))?!?\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?"
)
SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!")
NORMALIZE_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d{1,7}")


def categorize_formula(formula: str) -> list[str]:
    cats = [name for name, pat in FORMULA_CATEGORY_RULES if pat.search(formula)]
    return cats or ["OTHER"]


def referenced_sheets(formula: str, current_sheet: str) -> list[str]:
    names = set()
    for m in SHEET_REF_RE.finditer(formula):
        name = m.group(1) or m.group(2)
        if name and name != current_sheet:
            names.add(name)
    return sorted(names)


def normalize_pattern(formula: str) -> str:
    """Strip sheet-qualified and bare cell references to find repeated
    formula shapes (e.g. '=B2*C2' and '=B3*C3' both normalize the same)."""
    no_sheet_refs = re.sub(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!", "", formula)
    return NORMALIZE_REF_RE.sub("REF", no_sheet_refs)


@dataclass
class SheetInventory:
    name: str
    index: int
    sheet_state: str
    dimensions: str
    max_row: int
    max_col: int
    formula_count: int
    merged_ranges: list = field(default_factory=list)
    table_names: list = field(default_factory=list)
    autofilter_ref: str | None = None
    protection_enabled: bool = False
    protection_password_hash_present: bool = False
    print_area: str | None = None
    freeze_panes: str | None = None
    sheet_view_tab_selected: bool = False
    comment_count: int = 0
    hyperlink_count: int = 0
    conditional_formatting_count: int = 0
    image_or_shape_present: bool = False


def load_workbook(path: Path):
    return openpyxl.load_workbook(path, keep_vba=True, data_only=False, read_only=False)


def scan_sheets(wb) -> list[SheetInventory]:
    out = []
    for idx, name in enumerate(wb.sheetnames):
        ws = wb[name]
        formula_count = 0
        comment_count = 0
        hyperlink_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    formula_count += 1
                if cell.comment is not None:
                    comment_count += 1
        hyperlink_count = len(ws._hyperlinks) if hasattr(ws, "_hyperlinks") else 0
        out.append(
            SheetInventory(
                name=name,
                index=idx,
                sheet_state=ws.sheet_state,
                dimensions=ws.dimensions,
                max_row=ws.max_row,
                max_col=ws.max_column,
                formula_count=formula_count,
                merged_ranges=[str(r) for r in ws.merged_cells.ranges],
                table_names=list(ws.tables.keys()) if hasattr(ws, "tables") else [],
                autofilter_ref=ws.auto_filter.ref,
                protection_enabled=bool(ws.protection.sheet),
                protection_password_hash_present=bool(ws.protection.password),
                print_area=ws.print_area or None,
                freeze_panes=ws.freeze_panes,
                comment_count=comment_count,
                hyperlink_count=hyperlink_count,
                conditional_formatting_count=len(list(ws.conditional_formatting)),
                image_or_shape_present=bool(getattr(ws, "_images", [])),
            )
        )
    return out


def scan_formulas(wb) -> list[dict]:
    catalog = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                is_formula = cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                )
                if not is_formula:
                    continue
                formula = cell.value
                array_formula = False
                if hasattr(formula, "text"):  # ArrayFormula object
                    array_formula = True
                    formula = formula.text
                catalog.append(
                    {
                        "sheet": name,
                        "cell": cell.coordinate,
                        "formula": formula,
                        "array_formula": array_formula,
                        "referenced_sheets": referenced_sheets(formula, name),
                        "category": categorize_formula(formula),
                        "pattern_id": normalize_pattern(formula),
                    }
                )
    return catalog


def scan_named_ranges(wb) -> list[dict]:
    out = []
    for name, dn in wb.defined_names.items():
        out.append(
            {
                "name": name,
                "scope": "workbook",
                "refers_to": dn.value,
                "is_ref_error": "#REF!" in (dn.value or ""),
            }
        )
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for name, dn in getattr(ws, "defined_names", {}).items():
            out.append(
                {
                    "name": name,
                    "scope": sheet_name,
                    "refers_to": dn.value,
                    "is_ref_error": "#REF!" in (dn.value or ""),
                }
            )
    return out
