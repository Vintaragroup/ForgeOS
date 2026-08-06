"""Populated-workbook importer: reads a recalculated job workbook's raw
line items, independently recomputes each section's totals from the
documented business rules (docs/business-rules.md Rules 1-2), and stages
both that recompute and the workbook's own authoritative totals for the
diff harness to compare.

Row locations for COST SUMMARY's two bands (category vs. component, see
Rule 5) are resolved dynamically from the Phase 0 formula catalog rather
than hardcoded, since row numbers are an accident of one workbook's
history, not a stable contract.
"""
from __future__ import annotations

import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from tools.workbook_audit.hashcheck import REPO_ROOT, sha256_of

FORMULA_CATALOG = REPO_ROOT / "artifacts" / "formula_catalog.json"

DEPARTMENT_ROWS = list(range(10, 26))  # COMPONENT sheet labor rows 10-25 (business-rules.md Rule 1)


def _formula_catalog() -> list[dict]:
    return json.loads(FORMULA_CATALOG.read_text())


def find_cost_summary_row(sheet_name: str, column: str) -> int:
    """Locate the COST SUMMARY row that represents `sheet_name`, by
    finding the formula in `column` that references it — e.g. column 'A'
    for component rows (A24='COMPONENT 1'!A5), column 'D' for category
    rows (D8=Flooring!J11)."""
    needle = f"'{sheet_name}'!" if " " in sheet_name else f"{sheet_name}!"
    for f in _formula_catalog():
        if f["sheet"] != "COST SUMMARY":
            continue
        cell_col = re.match(r"([A-Z]+)\d+", f["cell"]).group(1)
        if cell_col != column:
            continue
        if needle in f["formula"]:
            return int(re.match(r"[A-Z]+(\d+)", f["cell"]).group(1))
    raise LookupError(f"No COST SUMMARY row found referencing {sheet_name} in column {column}")


def import_component_section(wb, cs_wb, sheet_name: str) -> dict:
    """Read a COMPONENT sheet's raw material + labor lines, independently
    recompute totals per Rules 1-2, and pair with COST SUMMARY's own
    values for that row (component band)."""
    ws = wb[sheet_name]
    row = find_cost_summary_row(sheet_name, "A")
    cs = cs_wb["COST SUMMARY"]

    line_items = []
    material_total = 0.0
    for r in range(10, 60):
        qty, unit_cost = ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value
        if qty is None or unit_cost is None:
            continue
        total = qty * unit_cost
        material_total += total
        line_items.append({
            "line_type": "material", "row_number": r,
            "description": ws.cell(row=r, column=1).value,
            "qty": qty, "unit_cost": unit_cost, "total_cost": total,
        })

    labor_total = 0.0
    for r in DEPARTMENT_ROWS:
        hours = ws.cell(row=r, column=7).value   # G
        rate = ws.cell(row=r, column=8).value     # H
        if not hours:
            continue
        total = hours * rate
        labor_total += total
        line_items.append({
            "line_type": "labor", "row_number": r,
            "description": ws.cell(row=r, column=6).value,
            "qty": hours, "unit_cost": rate, "total_cost": total,
        })

    return {
        "section_type": "component",
        "sheet_name": sheet_name,
        "cost_summary_row": row,
        "quantity": cs[f"B{row}"].value,
        "line_items": line_items,
        # imported_*_total are PER-UNIT (as authored on the component sheet
        # itself, before the section's own quantity multiplier). COST
        # SUMMARY's E/Z columns are qty-extended (E<row>=B<row>*D<row>,
        # confirmed via formula_catalog.json) -- the diff harness multiplies
        # by quantity before comparing. See docs/business-rules.md Rule 5.
        "imported_material_total": material_total,
        "imported_labor_total": labor_total,
        "source_total_direct_material": cs[f"E{row}"].value,   # qty-extended, matches E24=B24*D24
        "source_total_labor_cost": cs[f"Z{row}"].value,        # already qty-extended (G24=COMPONENT!G10*B24)
        "source_total_overhead": cs[f"AB{row}"].value,
        "source_total_cost": cs[f"AD{row}"].value,
    }


def import_category_section(wb, cs_wb, sheet_name: str) -> dict:
    """Read a category sheet's DIRECT MATERIALS PER COMPONENT block (cols
    G-J -- the cost input COST SUMMARY!D<row> actually reads via J11; the
    separate B-E "RENTAL COMPONENT PRICES" block is a client-facing price
    list and is NOT a COST SUMMARY input, confirmed via formula_catalog.json)
    plus per-unit labor hours, recompute, and pair with COST SUMMARY's own
    values for that row (category band)."""
    ws = wb[sheet_name]
    row = find_cost_summary_row(sheet_name, "D")
    cs = cs_wb["COST SUMMARY"]
    quantity = cs[f"B{row}"].value or 0

    line_items = []
    material_total = 0.0
    for r in range(10, 12):  # DIRECT MATERIALS PER COMPONENT block: row 10 is the only data row in this template
        qty, unit_cost = ws.cell(row=r, column=8).value, ws.cell(row=r, column=9).value  # H, I
        if not qty or unit_cost is None:
            continue
        total = qty * unit_cost
        material_total += total
        line_items.append({
            "line_type": "material", "row_number": r,
            "description": ws.cell(row=r, column=7).value,   # G
            "qty": qty, "unit_cost": unit_cost, "total_cost": total,
        })

    # Labor: category sheet stores HOURS-PER-UNIT (col H); COST SUMMARY
    # multiplies by quantity and by the COMPONENT 1 master rate (Rule 1).
    labor_total = 0.0
    for r in range(16, 32):
        hours_per_unit = ws.cell(row=r, column=8).value  # H
        rate = ws.cell(row=r, column=9).value            # I (already = COMPONENT 1 master rate)
        if not hours_per_unit or rate is None:
            continue
        total = hours_per_unit * quantity * rate
        labor_total += total
        line_items.append({
            "line_type": "labor", "row_number": r,
            "description": ws.cell(row=r, column=7).value,
            "qty": hours_per_unit * quantity, "unit_cost": rate, "total_cost": total,
        })

    return {
        "section_type": "category",
        "sheet_name": sheet_name,
        "cost_summary_row": row,
        "quantity": quantity,
        "line_items": line_items,
        "imported_material_total": material_total,
        "imported_labor_total": labor_total,
        "source_total_direct_material": cs[f"E{row}"].value,
        "source_total_labor_cost": cs[f"Z{row}"].value,
        "source_total_overhead": cs[f"AB{row}"].value,
        "source_total_cost": cs[f"AD{row}"].value,
    }


def import_job(
    conn: sqlite3.Connection,
    source_path: Path,
    recalculated_path: Path,
    component_sheets: list[str],
    category_sheets: list[str],
    job_name: str,
    is_synthetic: bool,
) -> int:
    wb = openpyxl.load_workbook(recalculated_path, data_only=True)      # raw line-item inputs, as recalculated
    cs_wb = openpyxl.load_workbook(recalculated_path, data_only=True)   # COST SUMMARY authoritative values

    cur = conn.cursor()
    cur.execute(
        "INSERT INTO job (source_file, source_hash, recalculated_file, imported_at, job_name, is_synthetic) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (str(source_path), sha256_of(source_path), str(recalculated_path),
         datetime.now(timezone.utc).isoformat(), job_name, int(is_synthetic)),
    )
    job_id = cur.lastrowid

    sections = (
        [import_component_section(wb, cs_wb, s) for s in component_sheets]
        + [import_category_section(wb, cs_wb, s) for s in category_sheets]
    )

    for sec in sections:
        cur.execute(
            "INSERT INTO estimate_section "
            "(job_id, section_type, sheet_name, cost_summary_row, quantity, "
            " total_direct_material, total_labor_cost, total_overhead, total_cost) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (job_id, sec["section_type"], sec["sheet_name"], sec["cost_summary_row"], sec["quantity"],
             sec["source_total_direct_material"], sec["source_total_labor_cost"],
             sec["source_total_overhead"], sec["source_total_cost"]),
        )
        section_id = cur.lastrowid
        for li in sec["line_items"]:
            cur.execute(
                "INSERT INTO line_item (section_id, line_type, row_number, description, qty, unit_cost, total_cost) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (section_id, li["line_type"], li["row_number"], li["description"],
                 li["qty"], li["unit_cost"], li["total_cost"]),
            )
        # stash the importer's own independent recompute on the section dict for the diff harness
        sec["_section_id"] = section_id

    conn.commit()
    return job_id, sections
